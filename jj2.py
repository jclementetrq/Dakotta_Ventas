import pyodbc
import pandas as pd
import os
from datetime import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.styles import numbers
from openpyxl.styles import Alignment
import subprocess

# -------------------------
# Conexión a base de datos
# -------------------------
def conectar_db(dsn, user, password):
    conn_str = f'DSN=SAP10;UID=SYSTEM_INT;PWD=System_2025'
    return pyodbc.connect(conn_str)

# -------------------------
# Ejecutar query
# -------------------------
def ejecutar_query(conexion, query):
    return pd.read_sql(query, conexion)

# -------------------------
# Renombrar columnas mes_1, mes_2, ... a meses en español
# -------------------------
def renombrar_columnas_meses(df, prefijo="mes_", cantidad=4):
    mes_actual = datetime.now().month
    meses = [(mes_actual - (cantidad - i - 1)) % 12 or 12 for i in range(cantidad)]

    MESES_ES = [
        "", "enero", "febrero", "marzo", "abril", "mayo", "junio",
        "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre"
    ]
    nombres_meses = [MESES_ES[mes] for mes in meses]

    columnas_meses = [f"{prefijo}{i+1}" for i in range(cantidad)]
    mapeo = dict(zip(columnas_meses, nombres_meses))

    df = df.rename(columns=mapeo)
    return df

# -------------------------
# Guardar resultados en archivos separados por usuario
# -------------------------

def ajustar_ancho_columnas(archivo_excel, nombre_hoja, desde_columna_monedas=3):
    wb = load_workbook(archivo_excel)
    ws = wb[nombre_hoja]

    columnas_monedas = list(range(desde_columna_monedas, ws.max_column + 1))

    for idx, col in enumerate(ws.columns, start=1):
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        if idx in columnas_monedas:
            adjusted_width = max_length + 6  # espacio extra para moneda
        else:
            adjusted_width = max_length + 2

        ws.column_dimensions[col_letter].width = adjusted_width

    wb.save(archivo_excel)



def formatear_columnas_dolares(archivo_excel, nombre_hoja, desde_columna=3):
    wb = load_workbook(archivo_excel)
    ws = wb[nombre_hoja]

    for col_idx in range(desde_columna, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        for cell in ws[col_letter]:
            # Solo formatear las celdas que no son cabecera (asumimos fila 1 es cabecera)
            if cell.row != 1 and isinstance(cell.value, (int, float)):
                cell.number_format = '"$"#,##0.00'

    wb.save(archivo_excel)

def escribir_log(mensaje, archivo_log="C:\\Users\\Administrador\\OneDrive - TRACTOMAQ S.A\\Escritorio\\LOGS GIT JC\\LOGMEDIDOR1\\log.txt"):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with open(archivo_log, "a", encoding="utf-8") as f:
        f.write(f"[{timestamp}] {mensaje}\n")

def guardar_resultados_por_usuario(df, columna_usuario, nombre_hoja, carpeta_salida=
"C:\\Users\\Administrador\\OneDrive - TRACTOMAQ S.A\\Escritorio\\GITHUB JC\\data"
                                   ):
    if not os.path.exists(carpeta_salida):
        os.makedirs(carpeta_salida)

    usuarios = df[columna_usuario].unique()

    for usuario in usuarios:
        df_usuario = df[df[columna_usuario] == usuario]
        archivo_excel = os.path.join(carpeta_salida, f"{usuario}.xlsx")

        try:
            if os.path.exists(archivo_excel):
                with pd.ExcelWriter(archivo_excel, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                    df_usuario.to_excel(writer, sheet_name=nombre_hoja, index=False)
            else:
                with pd.ExcelWriter(archivo_excel, engine='openpyxl', mode='w') as writer:
                    df_usuario.to_excel(writer, sheet_name=nombre_hoja, index=False)

            ajustar_ancho_columnas(archivo_excel, nombre_hoja, desde_columna_monedas=3)
            formatear_columnas_dolares(archivo_excel, nombre_hoja, desde_columna=3)

            if nombre_hoja == "VENTAS POR GRUPO":
                agregar_indicador_positivos(archivo_excel, nombre_hoja, desde_columna=3)
            elif nombre_hoja == "VENTA MENSUAL":
                agregar_totales_columnas(archivo_excel, nombre_hoja, desde_columna=3)
            elif nombre_hoja == "CUMPLIMIENTO MENSUAL":
                formatear_cumplimiento_mensual(archivo_excel, nombre_hoja)

            mensaje_ok = f"Archivo generado correctamente: '{archivo_excel}' (hoja: {nombre_hoja})"
            print(f"✔ {mensaje_ok}")
            escribir_log(mensaje_ok)

        except PermissionError:
            mensaje_error = f"No se pudo guardar o modificar el archivo '{archivo_excel}'. Puede estar abierto por otro usuario."
            print(f"⚠ {mensaje_error}")
            escribir_log(mensaje_error)

def agregar_indicador_positivos(archivo_excel, nombre_hoja, desde_columna=3):
    wb = load_workbook(archivo_excel)
    ws = wb[nombre_hoja]

    total_filas = ws.max_row - 1  # asumiendo fila 1 es encabezado

    fila_indicador = ws.max_row + 1  # fila donde pondremos los indicadores

    for col_idx in range(desde_columna, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)

        # Contar valores > 0 en esa columna (filas 2 hasta max_row)
        valores = [cell.value for cell in ws[col_letter][1:]]  # omitir cabecera
        num_positivos = sum(1 for v in valores if isinstance(v, (int, float)) and v > 0)

        """

                # Calcular porcentaje
                porcentaje = num_positivos / total_filas if total_filas > 0 else 0

                # Escribir resultado como porcentaje con formato (por ejemplo 0.56 = 56%)
                celda = ws[f"{col_letter}{fila_indicador}"]
                celda.value = porcentaje
                celda.number_format = '0.00%'  # formato porcentaje

                """

        texto = f"{num_positivos} de {total_filas}"
        celda = ws[f"{col_letter}{fila_indicador}"]
        celda.value = texto
        celda.alignment = Alignment(horizontal="center")

    wb.save(archivo_excel)

def agregar_totales_columnas(archivo_excel, nombre_hoja, desde_columna=3):
    wb = load_workbook(archivo_excel)
    ws = wb[nombre_hoja]

    fila_total = ws.max_row + 1  # fila donde pondremos los totales

    for col_idx in range(desde_columna, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)

        # Sumar todos los valores numéricos en la columna (omitir encabezado)
        valores = [cell.value for cell in ws[col_letter][1:]]
        suma_total = sum(v for v in valores if isinstance(v, (int, float)))

        celda = ws[f"{col_letter}{fila_total}"]
        celda.value = suma_total
        celda.number_format = '"$"#,##0.00'
        celda.alignment = Alignment(horizontal="right")

    wb.save(archivo_excel)

def formatear_cumplimiento_mensual(archivo_excel, nombre_hoja):
    wb = load_workbook(archivo_excel)
    ws = wb[nombre_hoja]

    total_presupuesto = 0
    total_venta = 0
    total_por_cumplir = 0

    for row in ws.iter_rows(min_row=2):  # saltar encabezado
        for cell in row:
            if cell.column_letter == "C":  # PRESUPUESTO
                if isinstance(cell.value, (int, float)):
                    total_presupuesto += cell.value
                    cell.number_format = '"$"#,##0.00'
            elif cell.column_letter == "D":  # VENTA
                if isinstance(cell.value, (int, float)):
                    total_venta += cell.value
                    cell.number_format = '"$"#,##0.00'
            elif cell.column_letter == "E":  # POR CUMPLIR
                if isinstance(cell.value, (int, float)):
                    total_por_cumplir += cell.value
                    cell.number_format = '"$"#,##0.00'
            elif cell.column_letter == "F":  # CUMPLIMIENTO
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '0.00%'

    fila_total = ws.max_row + 1
    ws[f"B{fila_total}"] = "TOTAL"
    ws[f"B{fila_total}"].alignment = Alignment(horizontal="right")

    ws[f"C{fila_total}"] = total_presupuesto
    ws[f"C{fila_total}"].number_format = '"$"#,##0.00'

    ws[f"D{fila_total}"] = total_venta
    ws[f"D{fila_total}"].number_format = '"$"#,##0.00'

    ws[f"E{fila_total}"] = total_por_cumplir
    ws[f"E{fila_total}"].number_format = '"$"#,##0.00'

    cumplimiento_total = total_venta / total_presupuesto if total_presupuesto else 0
    ws[f"F{fila_total}"] = cumplimiento_total
    ws[f"F{fila_total}"].number_format = '0.00%'

    wb.save(archivo_excel)

def guardar_consolidado(df1, df2, df3, nombre_archivo, hoja1, hoja2, hoja3, desde_columna_monedas=3):
    if os.path.exists(nombre_archivo):
        os.remove(nombre_archivo)

    with pd.ExcelWriter(nombre_archivo, engine='openpyxl') as writer:
        df1.to_excel(writer, sheet_name=hoja1, index=False)
        df2.to_excel(writer, sheet_name=hoja2, index=False)
        df3.to_excel(writer, sheet_name=hoja3, index=False)

    for hoja in [hoja1, hoja2, hoja3]:
        ajustar_ancho_columnas(nombre_archivo, hoja, desde_columna_monedas)
        formatear_columnas_dolares(nombre_archivo, hoja, desde_columna_monedas)

    agregar_indicador_positivos(nombre_archivo, hoja1, desde_columna_monedas)
    agregar_totales_columnas(nombre_archivo, hoja2, desde_columna_monedas)
    formatear_cumplimiento_mensual(nombre_archivo, hoja3)

    print(f"✔ Consolidado generado: {nombre_archivo}")
    escribir_log(f"Consolidado generado: {nombre_archivo}")

# -------------------------
# MAIN
# -------------------------
if __name__ == "__main__":
    # Configura tus parámetros de conexión y queries
    dsn = 'tu_dsn'
    user = 'tu_usuario'
    password = 'tu_password'

    # Primer query (hoja 1)
    query1 = """SELECT * FROM DB_CATAECSA.VW_VENTA_GRUPO_CLIENTE_JC ORDER BY "ASESOR" ASC, "CLIENTE" ASC
    """

    # Segundo query (hoja 2 con columnas mes_1, mes_2, ...)
    query2 = """SELECT * FROM DB_CATAECSA.VW_VENTA_TRIMESTRAL_CLIENTE_JC ORDER BY "ASESOR" ASC, "CLIENTE" ASC

            """

    # Tercer query (hoja 3 - cumplimiento)
    query3 = """SELECT * FROM DB_CATAECSA.VW_VENTA_CUMPLIMIENTO_GRUPO_JC ORDER BY "ASESOR" ASC,"GRUPO" ASC
    """

    # Conectarse a la base
    conexion = conectar_db(dsn, user, password)

    # Hoja 1: VENTAS POR GRUPO
    df_hoja1 = ejecutar_query(conexion, query1)
    guardar_resultados_por_usuario(df_hoja1, columna_usuario="ASESOR", nombre_hoja="VENTAS POR GRUPO")

    # Hoja 2: VENTA MENSUAL
    df_hoja2 = ejecutar_query(conexion, query2)
    df_hoja2 = renombrar_columnas_meses(df_hoja2)


    # -----------------------------
    # Añadir presupuesto a clientes
    # -----------------------------
    df_presupuesto = pd.read_excel("C:\\Users\\Administrador\\OneDrive - TRACTOMAQ S.A\\Escritorio\\presupuestos_clientes.xlsx")

    # Unir con df_hoja2 por CLIENTE
    df_hoja2 = pd.merge(df_hoja2, df_presupuesto, on="CLIENTE", how="left")

    # Reemplazar valores faltantes por 0
    df_hoja2["PRESUPUESTO"] = df_hoja2["PRESUPUESTO"].fillna(0)

    # Mover 'PRESUPUESTO' al final
    cols = [col for col in df_hoja2.columns if col != "PRESUPUESTO"] + ["PRESUPUESTO"]
    df_hoja2 = df_hoja2[cols]

    # Guardar archivo por asesor
    guardar_resultados_por_usuario(df_hoja2, columna_usuario="ASESOR", nombre_hoja="VENTA MENSUAL")

    df_hoja3 = ejecutar_query(conexion, query3)
    guardar_resultados_por_usuario(df_hoja3, columna_usuario="ASESOR", nombre_hoja="CUMPLIMIENTO MENSUAL")

    # Archivo unificado
    nombre_consolidado = os.path.join(
        "C:\\Users\\Administrador\\OneDrive - TRACTOMAQ S.A\\Escritorio\\GITHUB JC\\data",
        "STROBEL CORDERO MARIA ELISABETH.xlsx"
    )
    guardar_consolidado(df_hoja1, df_hoja2, df_hoja3, nombre_consolidado,
                        hoja1="VENTAS POR GRUPO", hoja2="VENTA MENSUAL", hoja3="CUMPLIMIENTO MENSUAL")

    #agregar_totales_columnas(nombre_consolidado, "VENTA MENSUAL", desde_columna=3)

    print("✔ Archivos generados correctamente.")



    try:
        repo_dir = r"C:\Users\Administrador\OneDrive - TRACTOMAQ S.A\Escritorio\GITHUB JC"
        os.chdir(repo_dir)

        GIT_PATH = r"C:\Program Files\Git\bin\git.exe"  # o donde tengas Git instalado

        subprocess.run([GIT_PATH, "add", "data"], check=True)
        subprocess.run(
            [GIT_PATH, "commit", "-m", f"Actualización automática {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"],
            check=True)
        subprocess.run([GIT_PATH, "push", "origin", "main"], check=True)

        print("✅ Cambios subidos correctamente a GitHub.")

    except subprocess.CalledProcessError as e:
        print("❌ Error al hacer push a GitHub:", e)